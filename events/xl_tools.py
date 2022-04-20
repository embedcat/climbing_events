import operator
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from config import settings
from events import services
from events.models import Event, Participant


def load_template(filename: str) -> Workbook or None:
    try:
        return load_workbook(filename=filename)
    except FileNotFoundError:
        return None


def export_participants_to_start_list(event: Event):
    ROW_OFFSET = 8
    book = load_workbook(filename='static/events/xl_templates/startlist_template.xlsx')
    sheet = book.active
    sheet.cell(row=1, column=1).value = event.gym
    sheet.cell(row=2, column=1).value = event.title
    sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
    sheet.cell(row=3, column=6).value = event.date

    for set_no in range(event.set_num):
        sheet = book.copy_worksheet(book.worksheets[0])
        sheet.title = f'Сет {set_no + 1}'
        sheet.cell(row=5, column=1).value = sheet.title
        queryset = event.participant.filter(set_index=set_no)
        participants = sorted(queryset, key=operator.attrgetter('last_name'))
        for index, p in enumerate(participants):
            sheet.cell(row=ROW_OFFSET + index, column=1).value = index + 1
            sheet.cell(row=ROW_OFFSET + index, column=2).value = f'{p.last_name} {p.first_name}'
            sheet.cell(row=ROW_OFFSET + index, column=3).value = p.birth_year
            sheet.cell(row=ROW_OFFSET + index, column=4).value = p.get_gender_display()
            sheet.cell(row=ROW_OFFSET + index, column=5).value = p.city
            sheet.cell(row=ROW_OFFSET + index, column=6).value = p.get_grade_display()
            sheet.cell(row=ROW_OFFSET + index, column=7).value = p.team
            group_list = services.get_group_list(event=event)
            sheet.cell(row=ROW_OFFSET + index, column=8).value = group_list[p.group_index] if group_list != [] else ''
            sheet.cell(row=ROW_OFFSET + index, column=9).value = p.pin
    book.remove(book.worksheets[0])
    book.close()

    return save_virtual_workbook(book)


def export_result(event: Event):
    ROW_OFFSET = 9
    HEADS_ROW = 8
    book = load_workbook(filename='static/events/xl_templates/result_template.xlsx')
    sheet = book.active
    sheet.cell(row=1, column=1).value = event.gym
    sheet.cell(row=2, column=1).value = event.title
    sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
    sheet.cell(row=3, column=6).value = event.date

    result = services.get_results(event=event, full_results=True)
    routes = event.route.all().order_by('number')

    for gender in (Participant.GENDER_MALE, Participant.GENDER_FEMALE):
        gender_data = result[gender]
        for group_index, group in enumerate(gender_data):
            participants = group['data']
            scores = group['scores']
            title = f"{group['name']}_{gender}"
            sheet = book.copy_worksheet(book.worksheets[0])
            sheet.title = title
            for index, p in enumerate(participants):
                sheet.cell(row=ROW_OFFSET + index, column=1).value = index + 1
                sheet.cell(row=ROW_OFFSET + index,
                           column=2).value = f"{p['participant'].last_name} {p['participant'].first_name}"
                sheet.cell(row=ROW_OFFSET + index, column=3).value = p['participant'].birth_year
                sheet.cell(row=ROW_OFFSET + index, column=4).value = p['participant'].city
                sheet.cell(row=ROW_OFFSET + index, column=5).value = p['participant'].get_grade_display()
                sheet.cell(row=ROW_OFFSET + index, column=6).value = p['participant'].team
                sheet.cell(row=ROW_OFFSET + index, column=7).value = p['score']
                for num, accent in enumerate(p['accents']):
                    sheet.cell(row=HEADS_ROW, column=8 + num).value = f"T#{num + 1}"
                    sheet.cell(row=HEADS_ROW - 1, column=8 + num).value = routes[num].grade
                    sheet.cell(row=HEADS_ROW - 2, column=8 + num).value = scores[num].replace('\n', '/')
                    sheet.cell(row=ROW_OFFSET + index, column=8 + num).value = accent
                sheet.cell(row=HEADS_ROW, column=8 + len(p['accents'])).value = "Итог"
                sheet.cell(row=ROW_OFFSET + index, column=8 + len(p['accents'])).value = p['score']

    book.remove(book.worksheets[0])
    path = os.path.join(settings.PROTOCOLS_PATH, f'{event.id}')
    if not os.path.exists(path=path):
        os.mkdir(path)
    file = os.path.join(path, f"results_{datetime.today().strftime('%Y-%m-%d-%H%M%S')}.xlsx")
    book.save(file)
